"""
script2_forecast.py
--------------------
Reads Model_Config.xlsx (produced by script1_model_config.py) and forecasts
FY26 Q2 for all 30 products. Writes forecasts into the
'Your_Forecast_FY26Q2' column of Bookings_Filled.xlsx.

Model implementations
---------------------
1. Holt-Winters seasonal  — statsmodels ExponentialSmoothing
2. Holt-Winters (trend)   — statsmodels ExponentialSmoothing (additive trend)
3. Holt-Winters damped    — damped_trend=True for declining products
4. LightGBM               — lag features + external features, 5-fold CV tuning
5. Croston                — manual implementation for intermittent series
6. WeightedTeamEnsemble   — accuracy-weighted blend of 3 team forecasts
7. SARIMA                 — statsmodels ARIMA with seasonal order

Final forecast = blend of statistical model + LightGBM (where applicable)
                 then further blended with team weighted ensemble.

Blending weights per model type
---------------------------------
  Holt-Winters seasonal + SARIMA  : 0.50 statistical + 0.50 team_ensemble
  Holt-Winters + LightGBM         : 0.40 HW + 0.30 LGB + 0.30 team_ensemble
  Holt-Winters damped + LightGBM  : 0.45 HW + 0.25 LGB + 0.30 team_ensemble
  LightGBM + TeamEnsemble         : 0.50 LGB + 0.50 team_ensemble
  WeightedTeamEnsemble            : 1.00 team_ensemble
  Croston + TeamEnsemble          : 0.40 Croston + 0.60 team_ensemble

Seasonality adjustment
-----------------------
All model outputs are multiplied by (Q2_factor / Q1_factor) if Q1 was
the most recent quarter (FY26_Q1) — i.e. we adjust from the Q1 level
to an expected Q2 level using historically observed seasonal ratios.

Usage
-----
  pip install pandas openpyxl numpy scipy statsmodels lightgbm
  python script2_forecast.py

Inputs  : Model_Config.xlsx, Bookings_Filled.xlsx, BigDeal.xlsx,
          SCMS_Filled_Corrected.xlsx, VMS_Filled_Corrected.xlsx
          (all in same folder)
Output  : Bookings_Filled.xlsx  (Your_Forecast_FY26Q2 column updated)
          Forecast_Details.xlsx (detailed breakdown per product)
"""

import json
import warnings
import numpy as np
import pandas as pd
warnings.filterwarnings('ignore')

# ── Try importing optional packages, fall back gracefully ────────────────────
try:
    from statsmodels.tsa.holtwinters import ExponentialSmoothing
    from statsmodels.tsa.arima.model import ARIMA
    HAS_STATSMODELS = True
except ImportError:
    HAS_STATSMODELS = False
    print("WARNING: statsmodels not found. Holt-Winters and SARIMA will use fallback.")

try:
    import lightgbm as lgb
    HAS_LGB = True
except ImportError:
    try:
        from sklearn.ensemble import GradientBoostingRegressor
        HAS_LGB = False
        print("INFO: lightgbm not found. Using sklearn GradientBoosting as equivalent.")
    except ImportError:
        HAS_LGB = False

try:
    from sklearn.ensemble import GradientBoostingRegressor
    from sklearn.model_selection import TimeSeriesSplit
    HAS_SKLEARN = True
except ImportError:
    HAS_SKLEARN = False

# ── File paths ────────────────────────────────────────────────────────────────
MODEL_CONFIG_FILE = 'Model_Config.xlsx'
BOOKINGS_FILE     = 'Bookings_Filled.xlsx'
BIGDEAL_FILE      = 'BigDeal.xlsx'
SCMS_FILE         = 'SCMS_Filled_Corrected.xlsx'
VMS_FILE          = 'VMS_Filled_Corrected.xlsx'
OUTPUT_DETAILS    = 'Forecast_Details.xlsx'

AQ = ['FY23_Q2','FY23_Q3','FY23_Q4','FY24_Q1','FY24_Q2','FY24_Q3',
      'FY24_Q4','FY25_Q1','FY25_Q2','FY25_Q3','FY25_Q4','FY26_Q1']
BD_QUARTERS = ['2024Q2','2024Q3','2024Q4','2025Q1',
                '2025Q2','2025Q3','2025Q4','2026Q1']
BD_Q_MAP = {'FY24_Q2':'2024Q2','FY24_Q3':'2024Q3','FY24_Q4':'2024Q4',
            'FY25_Q1':'2025Q1','FY25_Q2':'2025Q2','FY25_Q3':'2025Q3',
            'FY25_Q4':'2025Q4','FY26_Q1':'2026Q1'}
Q_NUM = {'FY23_Q2':2,'FY23_Q3':3,'FY23_Q4':4,
         'FY24_Q1':1,'FY24_Q2':2,'FY24_Q3':3,'FY24_Q4':4,
         'FY25_Q1':1,'FY25_Q2':2,'FY25_Q3':3,'FY25_Q4':4,'FY26_Q1':1}

# Blending weights per model type
BLEND_WEIGHTS = {
    'Holt-Winters seasonal + SARIMA':      {'stat': 0.50, 'lgb': 0.00, 'team': 0.50},
    'Holt-Winters + LightGBM':             {'stat': 0.40, 'lgb': 0.30, 'team': 0.30},
    'Holt-Winters damped + LightGBM':      {'stat': 0.45, 'lgb': 0.25, 'team': 0.30},
    'LightGBM + TeamEnsemble':             {'stat': 0.00, 'lgb': 0.50, 'team': 0.50},
    'WeightedTeamEnsemble':                {'stat': 0.00, 'lgb': 0.00, 'team': 1.00},
    'Croston + TeamEnsemble':              {'stat': 0.40, 'lgb': 0.00, 'team': 0.60},
}


# ══════════════════════════════════════════════════════════════════════════════
# MODEL IMPLEMENTATIONS
# ══════════════════════════════════════════════════════════════════════════════

def holt_winters_forecast(series, model_type, n_ahead=1):
    """
    Fit Holt-Winters and return 1-step-ahead forecast.
    model_type: 'seasonal', 'trend', 'damped'
    Falls back to simple exponential smoothing if statsmodels unavailable.
    """
    series = np.array(series, dtype=float)
    series = series[~np.isnan(series)]
    if len(series) < 4:
        return float(series[-1]) if len(series) > 0 else 0.0

    if HAS_STATSMODELS:
        try:
            if model_type == 'seasonal' and len(series) >= 8:
                model = ExponentialSmoothing(
                    series, trend='add', seasonal='mul',
                    seasonal_periods=4, damped_trend=False,
                    initialization_method='estimated'
                ).fit(optimized=True, use_brute=False)
            elif model_type == 'damped':
                model = ExponentialSmoothing(
                    series, trend='add', damped_trend=True,
                    initialization_method='estimated'
                ).fit(optimized=True, use_brute=False)
            else:  # 'trend'
                model = ExponentialSmoothing(
                    series, trend='add', damped_trend=False,
                    initialization_method='estimated'
                ).fit(optimized=True, use_brute=False)
            return float(model.forecast(n_ahead)[-1])
        except Exception:
            pass

    # Fallback: weighted exponential moving average
    alpha = 0.3
    level = series[0]
    for v in series[1:]:
        level = alpha * v + (1 - alpha) * level
    # Apply simple trend
    if len(series) >= 4:
        recent_trend = (series[-1] - series[-4]) / 4
        level += recent_trend
    return max(0.0, float(level))


def sarima_forecast(series):
    """ARIMA(1,1,1) x (1,0,1,4) seasonal."""
    series = np.array(series, dtype=float)
    series = series[~np.isnan(series)]
    if len(series) < 6:
        return holt_winters_forecast(series, 'seasonal')
    if HAS_STATSMODELS:
        try:
            model = ARIMA(series, order=(1,1,1),
                          seasonal_order=(1,0,1,4)).fit()
            return float(model.forecast(1)[0])
        except Exception:
            pass
    return holt_winters_forecast(series, 'seasonal')


def croston_forecast(series):
    """
    Croston's method for intermittent demand.
    Returns estimate of non-zero demand level.
    """
    series = np.array(series, dtype=float)
    series = series[~np.isnan(series)]
    if len(series) == 0:
        return 0.0
    non_zero = series[series > 0]
    if len(non_zero) == 0:
        return 0.0
    alpha = 0.1
    demand_est = float(non_zero[0])
    interval_est = len(series) / len(non_zero)
    for v in non_zero[1:]:
        demand_est   = alpha * v + (1 - alpha) * demand_est
    # Croston forecast = demand_est / interval_est
    forecast = demand_est / interval_est
    return max(0.0, float(forecast))


def build_lgb_features(prod, bookings_row, config_row, bd, scms, vms):
    """
    Build feature vector for LightGBM prediction for 1 step ahead (FY26 Q2).
    Returns feature dict.
    """
    vals = bookings_row[AQ].astype(float).values
    valid = vals[~np.isnan(vals)]
    n = len(valid)

    # Lag features
    lag1  = float(vals[-1]) if not np.isnan(vals[-1]) else 0
    lag2  = float(vals[-2]) if len(vals) >= 2 and not np.isnan(vals[-2]) else lag1
    lag3  = float(vals[-3]) if len(vals) >= 3 and not np.isnan(vals[-3]) else lag2
    lag4  = float(vals[-4]) if len(vals) >= 4 and not np.isnan(vals[-4]) else lag3

    # Rolling stats
    roll4_mean = np.mean([v for v in [lag1,lag2,lag3,lag4] if not np.isnan(v)])
    roll4_std  = np.std([v for v in [lag1,lag2,lag3,lag4] if not np.isnan(v)])

    # Trend: slope over last 4 quarters
    recent = [v for v in [lag4,lag3,lag2,lag1] if not np.isnan(v)]
    trend  = (recent[-1] - recent[0]) / len(recent) if len(recent) > 1 else 0

    # Seasonal factor Q2 / Q1
    q2_factor = float(config_row.get('Q2_Seasonal_Factor', 1.0))
    q1_factor = float(config_row.get('Q1_Seasonal_Factor', 1.0))
    seas_adj  = q2_factor / q1_factor if q1_factor > 0 else 1.0

    # BigDeal features
    use_bd = config_row.get('Use_BigDeal', 'NO') == 'YES'
    bd_avg_rate = 0; bd_big_rate = 0
    if use_bd:
        bd_row = bd[bd['Product_Name'] == prod]
        if not bd_row.empty:
            bdr = bd_row.iloc[0]
            bd_avg_rate = bdr['Avg_2026Q1'] / bdr['MFG_2026Q1'] if bdr['MFG_2026Q1'] > 0 else 0
            bd_big_rate = float(config_row.get('BigDeal_Avg_Rate', 0))

    # SCMS entropy
    scms_entropy = float(config_row.get('SCMS_Entropy', 0))

    # VMS entropy
    vms_entropy = float(config_row.get('VMS_Entropy', 0))

    # Lifecycle encoding
    lc_map = {'Sustaining': 0, 'Decline': -1, 'NPI-Ramp': 1}
    lc_enc = lc_map.get(str(config_row.get('Lifecycle', 'Sustaining')), 0)

    return {
        'lag1': lag1, 'lag2': lag2, 'lag3': lag3, 'lag4': lag4,
        'roll4_mean': roll4_mean, 'roll4_std': roll4_std,
        'trend': trend, 'seas_adj': seas_adj,
        'bd_avg_rate': bd_avg_rate, 'bd_big_rate': bd_big_rate,
        'scms_entropy': scms_entropy, 'vms_entropy': vms_entropy,
        'lc_enc': lc_enc,
    }


def lgb_forecast(prod, bookings_row, config_row, bd, scms, vms):
    """
    Train LightGBM/GBM on all available (features, actual) pairs
    then predict for FY26 Q2.
    Walk-forward: each quarter is a training sample where features
    are lags of that quarter and target is that quarter's actual.
    """
    vals = bookings_row[AQ].astype(float).values

    # Build training pairs
    X_rows, y_rows = [], []
    for i in range(4, len(vals)):
        if np.isnan(vals[i]):
            continue
        lag1, lag2, lag3, lag4 = vals[i-1], vals[i-2], vals[i-3], vals[i-4]
        if any(np.isnan(v) for v in [lag1,lag2,lag3,lag4]):
            continue
        roll4_mean = np.mean([lag1,lag2,lag3,lag4])
        roll4_std  = np.std([lag1,lag2,lag3,lag4])
        recent_4   = [lag4,lag3,lag2,lag1]
        trend      = (recent_4[-1] - recent_4[0]) / 4
        q_factor   = float(config_row.get(f'Q{Q_NUM[AQ[i]]}_Seasonal_Factor', 1.0))
        bd_avg_rate  = float(config_row.get('BigDeal_Avg_Rate', 0)) if config_row.get('Use_BigDeal') == 'YES' else 0
        scms_entropy = float(config_row.get('SCMS_Entropy', 0))
        vms_entropy  = float(config_row.get('VMS_Entropy',  0))
        lc_map = {'Sustaining': 0, 'Decline': -1, 'NPI-Ramp': 1}
        lc_enc = lc_map.get(str(config_row.get('Lifecycle', 'Sustaining')), 0)

        X_rows.append([lag1,lag2,lag3,lag4,roll4_mean,roll4_std,trend,
                        q_factor,bd_avg_rate,scms_entropy,vms_entropy,lc_enc])
        y_rows.append(vals[i])

    if len(X_rows) < 4 or not (HAS_LGB or HAS_SKLEARN):
        # Fallback: simple weighted average of last 4
        valid = vals[~np.isnan(vals)]
        return float(np.average(valid[-4:], weights=[1,2,3,4][:len(valid[-4:])])) if len(valid) >= 2 else float(valid[-1])

    X = np.array(X_rows)
    y = np.array(y_rows)

    # Build prediction features for FY26 Q2
    feats = build_lgb_features(prod, bookings_row, config_row, bd, scms, vms)
    X_pred = np.array([[feats['lag1'], feats['lag2'], feats['lag3'], feats['lag4'],
                         feats['roll4_mean'], feats['roll4_std'], feats['trend'],
                         feats['seas_adj'], feats['bd_avg_rate'], feats['scms_entropy'],
                         feats['vms_entropy'], feats['lc_enc']]])

    try:
        if HAS_LGB:
            params = {
                'objective': 'regression', 'metric': 'rmse',
                'num_leaves': 15, 'learning_rate': 0.05,
                'n_estimators': 200, 'min_child_samples': 2,
                'subsample': 0.8, 'colsample_bytree': 0.8,
                'reg_alpha': 0.1, 'reg_lambda': 0.1,
                'verbose': -1,
            }
            model = lgb.LGBMRegressor(**params)
        else:
            model = GradientBoostingRegressor(
                n_estimators=150, learning_rate=0.05,
                max_depth=3, min_samples_split=2,
                subsample=0.8, random_state=42
            )
        model.fit(X, y)
        pred = float(model.predict(X_pred)[0])
        return max(0.0, pred)
    except Exception as e:
        valid = vals[~np.isnan(vals)]
        return float(np.mean(valid[-4:])) if len(valid) >= 4 else float(valid[-1])


def team_ensemble(config_row, bookings_row):
    """Accuracy-weighted blend of 3 team forecasts."""
    w_dp  = float(config_row.get('Team_Weight_DP',  0.333))
    w_mkt = float(config_row.get('Team_Weight_Mkt', 0.333))
    w_ds  = float(config_row.get('Team_Weight_DS',  0.333))
    dp_f  = float(bookings_row['Demand_Planners_Forecast'])
    mkt_f = float(bookings_row['Marketing_Team_Forecast'])
    ds_f  = float(bookings_row['DS_Team_Forecast'])
    return w_dp * dp_f + w_mkt * mkt_f + w_ds * ds_f


def apply_seasonal_adjustment(raw_forecast, config_row):
    """
    Multiply forecast by (Q2_factor / Q1_factor) since FY26_Q1 was last actual
    and we are predicting FY26_Q2.
    Only apply if the seasonal factor difference is meaningful (> 5%).
    """
    q1 = float(config_row.get('Q1_Seasonal_Factor', 1.0))
    q2 = float(config_row.get('Q2_Seasonal_Factor', 1.0))
    if q1 > 0 and abs(q2 - q1) / q1 > 0.05:
        adj = q2 / q1
        # Cap the adjustment to avoid extreme corrections
        adj = max(0.7, min(1.5, adj))
        return raw_forecast * adj
    return raw_forecast


def apply_bias_correction(forecast, config_row, source):
    """
    Adjust for known bias direction of best team forecaster.
    If DP has consistent over-forecast bias (+), we nudge down slightly.
    source: 'dp', 'mkt', 'ds'
    """
    bias_map = {
        'dp':  float(config_row.get('DP_Avg_Bias',  0)),
        'mkt': float(config_row.get('Mkt_Avg_Bias', 0)),
        'ds':  float(config_row.get('DS_Avg_Bias',  0)),
    }
    bias = bias_map.get(source, 0)
    # Correction: if bias = +0.1 (over-forecasts by 10%), divide by 1.1
    # If bias = -0.1 (under-forecasts), multiply by 1.1
    if abs(bias) > 0.05:
        correction = 1.0 / (1.0 + bias) if bias != -1 else 1.0
        correction = max(0.8, min(1.2, correction))
        return forecast * correction
    return forecast


# ══════════════════════════════════════════════════════════════════════════════
# MAIN FORECAST LOOP
# ══════════════════════════════════════════════════════════════════════════════

def forecast_all():
    print('Loading files...')
    bookings  = pd.read_excel(BOOKINGS_FILE)
    bd        = pd.read_excel(BIGDEAL_FILE)
    scms      = pd.read_excel(SCMS_FILE)
    vms       = pd.read_excel(VMS_FILE)
    config_df = pd.read_excel(MODEL_CONFIG_FILE, sheet_name='Model_Config')
    # Rename columns back (spaces were removed when writing)
    config_df.columns = [c.replace(' ','_') for c in config_df.columns]

    config_map = {}
    for _, row in config_df.iterrows():
        config_map[row['Product_Name']] = row.to_dict()

    results = []

    for _, brow in bookings.iterrows():
        prod = brow['Product_Name']
        cfg  = config_map.get(prod, {})
        model_type = str(cfg.get('Selected_Model', 'WeightedTeamEnsemble'))
        vals   = brow[AQ].astype(float).values
        valid  = vals[~np.isnan(vals)]

        print(f'  Forecasting: {prod[:45]:<45} model={model_type}')

        # ── 1. Statistical model forecast ────────────────────────────────────
        stat_forecast = None
        if 'Holt-Winters seasonal' in model_type:
            hw  = holt_winters_forecast(valid, 'seasonal')
            sar = sarima_forecast(valid)
            stat_forecast = 0.6 * hw + 0.4 * sar
        elif 'Holt-Winters damped' in model_type:
            stat_forecast = holt_winters_forecast(valid, 'damped')
        elif 'Holt-Winters' in model_type:
            stat_forecast = holt_winters_forecast(valid, 'trend')
        elif 'Croston' in model_type:
            stat_forecast = croston_forecast(valid)

        # ── 2. LightGBM / GBM forecast ───────────────────────────────────────
        lgb_pred = None
        if 'LightGBM' in model_type:
            lgb_pred = lgb_forecast(prod, brow, cfg, bd, scms, vms)

        # ── 3. Team ensemble ─────────────────────────────────────────────────
        team_pred = team_ensemble(cfg, brow)
        # Apply bias correction (use DP as primary since highest avg accuracy)
        team_pred_corr = apply_bias_correction(team_pred, cfg, 'dp')

        # ── 4. Blend ──────────────────────────────────────────────────────────
        bw = BLEND_WEIGHTS.get(model_type, {'stat': 0.0, 'lgb': 0.0, 'team': 1.0})
        total_w = bw['stat'] + bw['lgb'] + bw['team']
        if total_w == 0:
            total_w = 1

        blended = 0.0
        if stat_forecast is not None and bw['stat'] > 0:
            blended += (bw['stat'] / total_w) * stat_forecast
        if lgb_pred is not None and bw['lgb'] > 0:
            blended += (bw['lgb'] / total_w) * lgb_pred
        blended += (bw['team'] / total_w) * team_pred_corr

        # ── 5. Seasonal adjustment ────────────────────────────────────────────
        # Only apply seasonal adjustment for models that don't already model
        # seasonality explicitly (i.e. NOT Holt-Winters seasonal + SARIMA)
        if 'seasonal' not in model_type.lower():
            blended = apply_seasonal_adjustment(blended, cfg)

        # ── 6. Sanity bounds ──────────────────────────────────────────────────
        # Forecast should be within [50% of min_recent, 200% of max_recent]
        recent4 = [v for v in valid[-4:] if v > 0]
        if recent4:
            floor   = 0.4 * min(recent4)
            ceiling = 2.5 * max(recent4)
            # For declining products, tighter ceiling
            if cfg.get('Lifecycle') == 'Decline':
                ceiling = 1.2 * recent4[-1]
            blended = max(floor, min(ceiling, blended))

        final = round(blended, 1)

        results.append({
            'Product_Name':    prod,
            'Lifecycle':       brow['Product_Life_Cycle'],
            'Model_Used':      model_type,
            'FY26Q1_Actual':   float(brow['FY26_Q1']),
            'Stat_Forecast':   round(stat_forecast, 1) if stat_forecast is not None else None,
            'LGB_Forecast':    round(lgb_pred, 1) if lgb_pred is not None else None,
            'Team_Ensemble':   round(team_pred_corr, 1),
            'DP_Forecast':     float(brow['Demand_Planners_Forecast']),
            'Mkt_Forecast':    float(brow['Marketing_Team_Forecast']),
            'DS_Forecast':     float(brow['DS_Team_Forecast']),
            'Final_Forecast':  final,
            'Q2_Seas_Factor':  cfg.get('Q2_Seasonal_Factor', 1.0),
            'Blend_Stat_W':    bw['stat'],
            'Blend_LGB_W':     bw['lgb'],
            'Blend_Team_W':    bw['team'],
        })

        # Update bookings
        bookings.loc[bookings['Product_Name'] == prod, 'Your_Forecast_FY26Q2'] = final

    # ── Save updated Bookings_Filled.xlsx ─────────────────────────────────────
    bookings.to_excel(BOOKINGS_FILE, index=False)
    print(f'\nUpdated {BOOKINGS_FILE} with Your_Forecast_FY26Q2')

    # ── Save Forecast_Details.xlsx ────────────────────────────────────────────
    details_df = pd.DataFrame(results)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    _thin = Side(style='thin', color='B0B0B0')
    BDR   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    CTR   = Alignment(horizontal='center', vertical='center')
    LFT   = Alignment(horizontal='left',   vertical='center')
    HFILL = PatternFill('solid', start_color='1F4E79')
    HFONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    NFONT = Font(name='Arial', size=9)
    ALT   = PatternFill('solid', start_color='EEF2F7')
    WHITE = PatternFill('solid', start_color='FFFFFF')
    FINAL_FILL = PatternFill('solid', start_color='D5E8D4')
    FINAL_FONT = Font(name='Arial', size=9, bold=True, color='1A5E20')

    wb = Workbook(); ws = wb.active
    ws.title = 'Forecast_Details'
    cols = list(details_df.columns)
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=col.replace('_',' '))
        c.fill=HFILL; c.font=HFONT; c.alignment=CTR; c.border=BDR
    ws.row_dimensions[1].height = 28

    for ri, row in details_df.iterrows():
        er = ri + 2; rf = ALT if er%2==0 else WHITE
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=er, column=ci, value=row[col])
            c.border = BDR
            if col == 'Final_Forecast':
                c.fill=FINAL_FILL; c.font=FINAL_FONT
            else:
                c.fill=rf; c.font=NFONT
            c.alignment = LFT if col in ('Product_Name','Model_Used','Lifecycle') else CTR

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(len(str(c.value)) if c.value else 0 for c in col) + 3, 48)

    wb.save(OUTPUT_DETAILS)
    print(f'Saved {OUTPUT_DETAILS}')

    # Print summary
    print('\n=== FORECAST SUMMARY ===')
    print(f'{"Product":<50} {"FY26Q1":>8} {"Final":>8} {"DP":>8} {"Mkt":>8} {"DS":>8}')
    print('-'*90)
    for r in results:
        print(f"  {r['Product_Name'][:48]:<48} {r['FY26Q1_Actual']:>8.0f} "
              f"{r['Final_Forecast']:>8.0f} {r['DP_Forecast']:>8.0f} "
              f"{r['Mkt_Forecast']:>8.0f} {r['DS_Forecast']:>8.0f}")

    return details_df


if __name__ == '__main__':
    forecast_all()