"""
script1_model_config.py
-----------------------
Analyses all product time series and writes a model configuration Excel file
containing the selected model, feature flags, seasonal factors, and all
metadata needed by script2_forecast.py.

Also writes a "Feature_Store" sheet with pre-computed derived features
(SCMS entropy, VMS dominant vertical, BigDeal avg rate, team accuracy weights)
so the forecast script can load them directly without re-computation.

Usage
-----
  pip install pandas openpyxl numpy scipy
  python script1_model_config.py

Inputs  : Bookings_Filled.xlsx, BigDeal.xlsx,
          SCMS_Filled_Corrected.xlsx, VMS_Filled_Corrected.xlsx,
          Forecast_Accuracy.xlsx
          (all in same folder)
Output  : Model_Config.xlsx
"""

import json
import numpy as np
import pandas as pd
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── File paths ────────────────────────────────────────────────────────────────
BOOKINGS_FILE = 'Bookings_Filled.xlsx'
BIGDEAL_FILE  = 'BigDeal.xlsx'
SCMS_FILE     = 'SCMS_Filled_Corrected.xlsx'
VMS_FILE      = 'VMS_Filled_Corrected.xlsx'
FA_FILE       = 'Forecast_Accuracy.xlsx'
OUTPUT_FILE   = 'Model_Config.xlsx'

# ── Constants ─────────────────────────────────────────────────────────────────
AQ = ['FY23_Q2','FY23_Q3','FY23_Q4','FY24_Q1','FY24_Q2','FY24_Q3',
      'FY24_Q4','FY25_Q1','FY25_Q2','FY25_Q3','FY25_Q4','FY26_Q1']

BD_QUARTERS = ['2024Q2','2024Q3','2024Q4','2025Q1',
                '2025Q2','2025Q3','2025Q4','2026Q1']
OVERLAP_Q   = ['FY24_Q2','FY24_Q3','FY24_Q4','FY25_Q1',
               'FY25_Q2','FY25_Q3','FY25_Q4','FY26_Q1']
BD_Q_MAP    = dict(zip(OVERLAP_Q, BD_QUARTERS))

# Cisco fiscal quarter numbers (Q1=Aug-Oct, Q2=Nov-Jan, Q3=Feb-Apr, Q4=May-Jul)
Q_NUM = {'FY23_Q2':2,'FY23_Q3':3,'FY23_Q4':4,
         'FY24_Q1':1,'FY24_Q2':2,'FY24_Q3':3,'FY24_Q4':4,
         'FY25_Q1':1,'FY25_Q2':2,'FY25_Q3':3,'FY25_Q4':4,'FY26_Q1':1}

# Product type mapping (from description analysis)
PROD_TYPES = {
    'Campus & Enterprise Networking': [
        'SWITCH Enterprise 48-Port UPOE','SWITCH Enterprise 24-Port PoE+',
        'SWITCH Enterprise 24-Port UPOE','SWITCH Enterprise 48-Port Fiber (Non-PoE)',
        'SWITCH Enterprise 25G Fiber','SWITCH Enterprise 48-Port PoE+ with 10G Uplink',
        'SWITCH Enterprise 24-Port PoE+ Compact','SWITCH Core 25G/100G Fiber',
        'SWITCH Core 100G Fiber','SWITCH Core 100G/400G Fiber',
        'WIRELESS ACCESS POINT WiFi6E (Internal Antenna) Indoor',
        'WIRELESS ACCESS POINT WiFi6 (Internal Antenna) Indoor',
        'WIRELESS ACCESS POINT WiFi6E (External Antenna) Outdoor',
        'WIRELESS ACCESS POINT WiFi6 (External Antenna) Outdoor',
        'WIRELESS ACCESS POINT WiFi6 (External Antenna) Indoor',
    ],
    'Data Center Infrastructure': [
        'SWITCH Data Center 25G/100G Leaf','SWITCH Data Center 100G Spine',
        'ROUTER Core Modular Chassis',
    ],
    'Industrial Networking': [
        'SWITCH Industrial 8-Port PoE','SWITCH Industrial 8-Port PoE Compact',
        'SWITCH Industrial Managed','SWITCH Industrial 24-Port PoE',
    ],
    'Routing, Security & WAN': [
        'ROUTER Enterprise Edge 10G/40G','ROUTER Branch 4-Port PoE',
        'ROUTER Branch 8-Port PoE','ROUTER Edge Aggregation Fiber',
        'SECURITY FIREWALL Next-Generation_1','SECURITY FIREWALL Next-Generation_2',
    ],
    'Collaboration & Voice': [
        'IP PHONE Enterprise Desk','IP CONFERENCE PHONE',
    ],
}
PROD_TO_TYPE = {p: t for t, prods in PROD_TYPES.items() for p in prods}

# External seasonality notes (Cisco fiscal calendar)
# Q1 = Aug-Oct: back-to-school/fiscal start → variable
# Q2 = Nov-Jan: holiday slowdown in some sectors but strong enterprise budget flush
# Q3 = Feb-Apr: spring ramp, budget deployment
# Q4 = May-Jul: fiscal year-end flush
EXTERNAL_CONTEXT = {
    'Q1': 'Aug-Oct: new FY budgets, back-to-school, variable demand',
    'Q2': 'Nov-Jan: holiday season, enterprise budget flush, strong for some products',
    'Q3': 'Feb-Apr: spring ramp, steady enterprise procurement',
    'Q4': 'May-Jul: fiscal year-end push, often peak quarter',
}

# Correlation insight per product type
CORRELATION_NOTES = {
    'Campus & Enterprise Networking': 'High correlation with education/government budget cycles (Q3-Q4). SCMS entropy > 1.8 indicates diversified demand = more predictable.',
    'Data Center Infrastructure':     'Driven by large DC build-out projects. BigDeal highly correlated with spikes. Avg demand (non-big-deal) is the stable baseline.',
    'Industrial Networking':          'Tied to manufacturing capex and utility infrastructure cycles. Energy/Utilities + Manufacturing dominate. Seasonal lag vs consumer products.',
    'Routing, Security & WAN':        'Security and routing driven by compliance cycles (Q4 budget) and WAN refresh. Service Provider buys independently of enterprise cycles.',
    'Collaboration & Voice':          'IP Phone in structural decline; conference phones stable. Enterprise budget flush (Q4) drives peaks. BigDeal = campus/HQ rollout events.',
}

# ── Style helpers ─────────────────────────────────────────────────────────────
_thin  = Side(style='thin', color='B0B0B0')
BDR    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CTR    = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
HFILL  = PatternFill('solid', start_color='1F4E79')
HFONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
NFONT  = Font(name='Arial', size=9)
ALT    = PatternFill('solid', start_color='EEF2F7')
WHITE  = PatternFill('solid', start_color='FFFFFF')
GREENFILL = PatternFill('solid', start_color='E2EFDA')
AMFILL    = PatternFill('solid', start_color='FFF2CC')
REDFILL   = PatternFill('solid', start_color='FFE0B2')


def _style_row(ws, row_idx, n_cols, is_header=False, fill=None):
    rf = fill or (ALT if row_idx % 2 == 0 else WHITE)
    for ci in range(1, n_cols + 1):
        c = ws.cell(row=row_idx, column=ci)
        c.border = BDR
        if is_header:
            c.fill = HFILL; c.font = HFONT; c.alignment = CTR
        else:
            c.fill = rf; c.font = NFONT; c.alignment = CTR


def _autowidth(ws, max_w=55):
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, max_w)


# ══════════════════════════════════════════════════════════════════════════════
# ANALYSIS FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def compute_seasonality(bookings):
    """Return dict: product → {1: factor, 2: factor, 3: factor, 4: factor}"""
    result = {}
    for _, row in bookings.iterrows():
        by_q = {1:[], 2:[], 3:[], 4:[]}
        for col in AQ:
            v = float(row[col]) if not np.isnan(float(row[col])) else np.nan
            if not np.isnan(v) and v > 0:
                by_q[Q_NUM[col]].append(v)
        overall = np.mean([v for lst in by_q.values() for v in lst])
        qf = {}
        for q, lst in by_q.items():
            qf[q] = round(np.mean(lst)/overall, 4) if lst and overall > 0 else 1.0
        result[row['Product_Name']] = qf
    return result


def compute_bd_relevance(bookings, bd):
    """Return dict: product → avg big deal fraction"""
    result = {}
    for _, row in bookings.iterrows():
        prod   = row['Product_Name']
        bd_row = bd[bd['Product_Name'] == prod]
        if bd_row.empty:
            result[prod] = 0.0; continue
        bdr   = bd_row.iloc[0]
        _BD_Q = {'FY24_Q2':'2024Q2','FY24_Q3':'2024Q3','FY24_Q4':'2024Q4',
                 'FY25_Q1':'2025Q1','FY25_Q2':'2025Q2','FY25_Q3':'2025Q3',
                 'FY25_Q4':'2025Q4','FY26_Q1':'2026Q1'}
        rates = [bdr[f'Big_{_BD_Q[q]}'] / bdr[f'MFG_{_BD_Q[q]}']
                 for q in _BD_Q
                 if bdr[f'MFG_{_BD_Q[q]}'] > 0 and bdr[f'Big_{_BD_Q[q]}'] > 0]
        result[prod] = round(np.mean(rates), 4) if rates else 0.0
    return result


def compute_scms_features(scms):
    """Return dict: product → {entropy, dominant_segment, dom_pct}"""
    result = {}
    for prod, grp in scms.groupby('Product_Name'):
        lt   = grp['2026Q1'].astype(float)
        tot  = lt.sum()
        if tot > 0:
            probs   = lt / tot
            p_pos   = probs[probs > 0].values
            entropy = -float(np.sum(p_pos * np.log2(p_pos)))
            dom_idx = lt.idxmax()
            dom_seg = grp.loc[dom_idx, 'Segment']
            dom_pct = round(lt.max() / tot, 4)
        else:
            entropy = 0.0; dom_seg = 'N/A'; dom_pct = 0.0
        result[prod] = {'entropy': round(entropy, 4),
                        'dominant_segment': dom_seg,
                        'dominant_pct': dom_pct}
    return result


def compute_vms_features(vms):
    """Return dict: product → {entropy, dominant_vertical, dom_pct, n_verticals}"""
    result = {}
    for prod, grp in vms.groupby('Product_Name'):
        lt   = grp['2026Q1'].astype(float)
        tot  = lt.sum()
        if tot > 0:
            probs   = lt / tot
            p_pos   = probs[probs > 0].values
            entropy = -float(np.sum(p_pos * np.log2(p_pos)))
            dom_idx = lt.idxmax()
            dom_vert = grp.loc[dom_idx, 'Vertical']
            dom_pct  = round(lt.max() / tot, 4)
        else:
            entropy = 0.0; dom_vert = 'N/A'; dom_pct = 0.0
        result[prod] = {'entropy': round(entropy, 4),
                        'dominant_vertical': dom_vert,
                        'dominant_pct': dom_pct,
                        'n_verticals': len(grp)}
    return result


def compute_team_weights(fa):
    """
    Compute accuracy-weighted ensemble weights per product.
    Weight = accuracy_i / sum(accuracies) across 3 quarters.
    """
    result = {}
    for _, row in fa.iterrows():
        prod = row['Product_Name']
        dp_acc  = np.mean([row['DP_FY26Q1_Acc'],  row['DP_FY25Q4_Acc'],  row['DP_FY25Q3_Acc']])
        mkt_acc = np.mean([row['Mkt_FY26Q1_Acc'], row['Mkt_FY25Q4_Acc'], row['Mkt_FY25Q3_Acc']])
        ds_acc  = np.mean([row['DS_FY26Q1_Acc'],  row['DS_FY25Q4_Acc'],  row['DS_FY25Q3_Acc']])
        # Bias correction factor: subtract abs(avg_bias) penalty
        dp_bias  = np.mean([abs(row['DP_FY26Q1_Bias']),  abs(row['DP_FY25Q4_Bias']),  abs(row['DP_FY25Q3_Bias'])])
        mkt_bias = np.mean([abs(row['Mkt_FY26Q1_Bias']), abs(row['Mkt_FY25Q4_Bias']), abs(row['Mkt_FY25Q3_Bias'])])
        ds_bias  = np.mean([abs(row['DS_FY26Q1_Bias']),  abs(row['DS_FY25Q4_Bias']),  abs(row['DS_FY25Q3_Bias'])])
        # Bias direction (positive = over-forecast, negative = under)
        dp_bias_dir  = np.mean([row['DP_FY26Q1_Bias'],  row['DP_FY25Q4_Bias'],  row['DP_FY25Q3_Bias']])
        mkt_bias_dir = np.mean([row['Mkt_FY26Q1_Bias'], row['Mkt_FY25Q4_Bias'], row['Mkt_FY25Q3_Bias']])
        ds_bias_dir  = np.mean([row['DS_FY26Q1_Bias'],  row['DS_FY25Q4_Bias'],  row['DS_FY25Q3_Bias']])
        # Adjusted score: accuracy penalised by bias magnitude
        dp_score  = dp_acc  - 0.3 * dp_bias
        mkt_score = mkt_acc - 0.3 * mkt_bias
        ds_score  = ds_acc  - 0.3 * ds_bias
        total     = dp_score + mkt_score + ds_score
        result[prod] = {
            'w_dp':  round(dp_score  / total, 4) if total > 0 else 0.333,
            'w_mkt': round(mkt_score / total, 4) if total > 0 else 0.333,
            'w_ds':  round(ds_score  / total, 4) if total > 0 else 0.333,
            'dp_acc':  round(dp_acc,  4), 'mkt_acc': round(mkt_acc, 4), 'ds_acc': round(ds_acc, 4),
            'dp_bias_dir': round(dp_bias_dir,4), 'mkt_bias_dir': round(mkt_bias_dir,4), 'ds_bias_dir': round(ds_bias_dir,4),
        }
    return result


def select_model(row, cv, slope_pct, r2, zero_frac, n, bd_pct, scms_ent):
    lc   = row['Product_Life_Cycle']
    prod = row['Product_Name']
    if lc == 'NPI-Ramp' or n < 6:
        return ('NPI / Short series',
                'WeightedTeamEnsemble',
                'Too few data points for own model; rely on team accuracy-weighted ensemble')
    if zero_frac > 0.25:
        return ('Intermittent',
                'Croston + TeamEnsemble',
                'Sparse demand pattern; Croston estimates non-zero mean, blended with team ensemble')
    if lc == 'Decline' and slope_pct < -0.015:
        return ('Trend — Declining',
                'Holt-Winters damped + LightGBM',
                'Clear decline; damped Holt-Winters prevents over-projection; LightGBM captures BD/SCMS non-linearity')
    if slope_pct > 0.04 and r2 > 0.3:
        return ('Trend — Growing',
                'Holt-Winters + LightGBM',
                'Strong growth; Holt-Winters captures level+trend; LightGBM adds BD/SCMS/VMS features')
    if cv < 0.25:
        return ('Stable — Seasonal',
                'Holt-Winters seasonal + SARIMA',
                'Low CV with Q1 dip pattern; multiplicative seasonality; SARIMA for autocorrelation')
    if cv < 0.45:
        return ('Stable — Moderate volatile',
                'Holt-Winters + LightGBM',
                'Moderate variance; Holt-Winters baseline + LightGBM for external feature capture')
    return ('Volatile',
            'LightGBM + TeamEnsemble',
            'High variance; LightGBM with lag features + team forecasts as blending features')


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print('Loading data...')
    bookings = pd.read_excel(BOOKINGS_FILE)
    bd       = pd.read_excel(BIGDEAL_FILE)
    scms     = pd.read_excel(SCMS_FILE)
    vms      = pd.read_excel(VMS_FILE)
    fa       = pd.read_excel(FA_FILE)

    print('Computing features...')
    seasonality   = compute_seasonality(bookings)
    bd_relevance  = compute_bd_relevance(bookings, bd)
    scms_features = compute_scms_features(scms)
    vms_features  = compute_vms_features(vms)
    team_weights  = compute_team_weights(fa)

    # ── Build config rows ─────────────────────────────────────────────────────
    config_rows = []
    for _, row in bookings.iterrows():
        prod  = row['Product_Name']
        lc    = row['Product_Life_Cycle']
        vals  = row[AQ].astype(float).values
        valid = vals[~np.isnan(vals) & (vals >= 0)]
        n     = len(valid)
        mean  = np.mean(valid) if n > 0 else 1
        std   = np.std(valid)  if n > 0 else 0
        cv    = std / mean if mean > 0 else 999
        if n >= 3:
            slope, _, r2, _, _ = stats.linregress(range(n), valid)
            slope_pct = slope / mean if mean > 0 else 0
        else:
            slope_pct = 0; r2 = 0
        zero_frac  = np.sum(valid == 0) / n if n > 0 else 0
        bd_pct     = bd_relevance.get(prod, 0)
        scms_info  = scms_features.get(prod, {})
        vms_info   = vms_features.get(prod, {})
        tw         = team_weights.get(prod, {})
        sfact      = seasonality.get(prod, {})
        scms_ent   = scms_info.get('entropy', 0)

        use_bd   = bd_pct > 0.08
        use_scms = scms_ent > 0.5 and prod in scms_features
        use_vms  = vms_info.get('n_verticals', 0) > 3

        ts_type, model, note = select_model(row, cv, slope_pct, r2, zero_frac, n, bd_pct, scms_ent)
        ptype = PROD_TO_TYPE.get(prod, 'Unknown')
        corr  = CORRELATION_NOTES.get(ptype, '')

        config_rows.append({
            'Product_Name':          prod,
            'Product_Life_Cycle':    lc,
            'Product_Type':          ptype,
            'TS_Classification':     ts_type,
            'CV':                    round(cv, 3),
            'Slope_Pct_Per_Q':       round(slope_pct, 4),
            'R2_Trend':              round(r2, 3),
            'Selected_Model':        model,
            'Model_Rationale':       note,
            'Use_BigDeal':           'YES' if use_bd else 'NO',
            'BigDeal_Avg_Rate':      round(bd_pct, 3),
            'Use_SCMS':              'YES' if use_scms else 'NO',
            'SCMS_Entropy':          round(scms_ent, 3),
            'SCMS_Dominant_Segment': scms_info.get('dominant_segment', 'N/A'),
            'Use_VMS':               'YES' if use_vms else 'NO',
            'VMS_Entropy':           round(vms_info.get('entropy', 0), 3),
            'VMS_Dominant_Vertical': vms_info.get('dominant_vertical', 'N/A'),
            'Q1_Seasonal_Factor':    sfact.get(1, 1.0),
            'Q2_Seasonal_Factor':    sfact.get(2, 1.0),
            'Q3_Seasonal_Factor':    sfact.get(3, 1.0),
            'Q4_Seasonal_Factor':    sfact.get(4, 1.0),
            'Team_Weight_DP':        tw.get('w_dp',  0.333),
            'Team_Weight_Mkt':       tw.get('w_mkt', 0.333),
            'Team_Weight_DS':        tw.get('w_ds',  0.333),
            'DP_Avg_Accuracy_3Q':    tw.get('dp_acc',  0),
            'Mkt_Avg_Accuracy_3Q':   tw.get('mkt_acc', 0),
            'DS_Avg_Accuracy_3Q':    tw.get('ds_acc',  0),
            'DP_Avg_Bias':           tw.get('dp_bias_dir',  0),
            'Mkt_Avg_Bias':          tw.get('mkt_bias_dir', 0),
            'DS_Avg_Bias':           tw.get('ds_bias_dir',  0),
            'Correlation_Context':   corr,
            'External_Q2_Context':   EXTERNAL_CONTEXT['Q2'],
        })

    config_df = pd.DataFrame(config_rows)

    # ── Build feature store ───────────────────────────────────────────────────
    feature_store_rows = []
    for prod in bookings['Product_Name']:
        row      = bookings[bookings['Product_Name'] == prod].iloc[0]
        bd_row   = bd[bd['Product_Name'] == prod]
        scms_grp = scms[scms['Product_Name'] == prod]
        vms_grp  = vms[vms['Product_Name'] == prod]
        tw       = team_weights.get(prod, {})

        # BigDeal Avg (non-big) for FY26Q1 as recent baseline
        bd_avg_2026q1 = 0
        if not bd_row.empty:
            bd_avg_2026q1 = bd_row.iloc[0]['Avg_2026Q1']

        # SCMS segment shares for FY26Q1 (as JSON string)
        scms_shares = {}
        if not scms_grp.empty:
            lt = scms_grp.set_index('Segment')['2026Q1'].astype(float)
            tot = lt.sum()
            if tot > 0:
                scms_shares = {k: round(v/tot, 4) for k,v in lt.items()}

        # VMS top 3 verticals for FY26Q1
        vms_top3 = {}
        if not vms_grp.empty:
            lt = vms_grp.set_index('Vertical')['2026Q1'].astype(float)
            tot = lt.sum()
            if tot > 0:
                vms_top3 = {k: round(v/tot, 4) for k,v in lt.nlargest(3).items()}

        feature_store_rows.append({
            'Product_Name':      prod,
            'FY26Q1_Actual':     row['FY26_Q1'],
            'FY25Q4_Actual':     row['FY25_Q4'],
            'FY25Q3_Actual':     row['FY25_Q3'],
            'FY25Q2_Actual':     row['FY25_Q2'],
            'FY25Q1_Actual':     row['FY25_Q1'],
            'BD_Avg_FY26Q1':     bd_avg_2026q1,
            'BD_BigRate_FY26Q1': bd_row.iloc[0]['Big_2026Q1'] / bd_row.iloc[0]['MFG_2026Q1']
                                 if not bd_row.empty and bd_row.iloc[0]['MFG_2026Q1'] > 0 else 0,
            'SCMS_Shares_2026Q1': json.dumps(scms_shares),
            'VMS_Top3_2026Q1':    json.dumps(vms_top3),
            'DP_Forecast_FY26Q2':  row['Demand_Planners_Forecast'],
            'Mkt_Forecast_FY26Q2': row['Marketing_Team_Forecast'],
            'DS_Forecast_FY26Q2':  row['DS_Team_Forecast'],
            'Team_Weighted_Ensemble': round(
                tw.get('w_dp',0.333) * row['Demand_Planners_Forecast'] +
                tw.get('w_mkt',0.333) * row['Marketing_Team_Forecast'] +
                tw.get('w_ds',0.333)  * row['DS_Team_Forecast'], 1),
        })

    feature_store_df = pd.DataFrame(feature_store_rows)

    # ── Write Excel ───────────────────────────────────────────────────────────
    print('Writing Model_Config.xlsx...')
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Model Config
    ws = wb.create_sheet('Model_Config')
    cols = list(config_df.columns)
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=col.replace('_',' '))
        c.fill=HFILL; c.font=HFONT; c.alignment=CTR; c.border=BDR
    ws.row_dimensions[1].height = 30

    for ri, row_data in config_df.iterrows():
        er   = ri + 2
        rf   = ALT if er % 2 == 0 else WHITE
        model_str = str(row_data.get('Selected_Model',''))
        for ci, col in enumerate(cols, 1):
            c   = ws.cell(row=er, column=ci, value=row_data[col])
            c.border = BDR; c.font = NFONT
            # Colour-code model column
            if col == 'Selected_Model':
                if 'LightGBM' in model_str:
                    c.fill = GREENFILL
                elif 'Holt-Winters' in model_str:
                    c.fill = AMFILL
                else:
                    c.fill = REDFILL
            elif col in ('Use_BigDeal','Use_SCMS','Use_VMS') and row_data[col]=='YES':
                c.fill = GREENFILL
            else:
                c.fill = rf
            c.alignment = CTR if col not in ('Product_Name','Model_Rationale',
                                              'Correlation_Context','External_Q2_Context',
                                              'Product_Type','TS_Classification') else LFT
    _autowidth(ws)

    # Sheet 2: Feature Store
    ws2 = wb.create_sheet('Feature_Store')
    cols2 = list(feature_store_df.columns)
    for ci, col in enumerate(cols2, 1):
        c = ws2.cell(row=1, column=ci, value=col.replace('_',' '))
        c.fill=HFILL; c.font=HFONT; c.alignment=CTR; c.border=BDR

    for ri, row_data in feature_store_df.iterrows():
        er = ri + 2; rf = ALT if er % 2 == 0 else WHITE
        for ci, col in enumerate(cols2, 1):
            c = ws2.cell(row=er, column=ci, value=row_data[col])
            c.fill=rf; c.font=NFONT; c.border=BDR; c.alignment=CTR
    _autowidth(ws2)

    # Sheet 3: Correlation & External Notes
    ws3 = wb.create_sheet('Insights')
    notes = [
        ('Topic', 'Insight'),
        ('Seasonality — Q1 dip',
         'Cisco fiscal Q1 (Aug-Oct) shows consistent dip for most products. FY26 Q2 (Nov-Jan) '
         'is a rebound quarter. Q2 seasonal factor > 1 for WiFi6 Indoor, IP Phone, IP Conference, '
         'Router Edge Aggregation — these are expected to be stronger than average.'),
        ('BigDeal → Bookings correlation',
         'Router Branch 4P/8P and Edge Aggregation show avg big-deal rate >40% — big deal volume '
         'is the primary driver, not avg deal volume. Industrial 24P (40%) and Industrial 8P Compact (35%) '
         'also big-deal driven. For these, forecasting Avg demand separately then adding BD estimate is optimal.'),
        ('SCMS entropy',
         'Switch Enterprise 48P UPOE has highest entropy (2.19) = demand spread across all segments. '
         'Switch Industrial Managed has near-zero entropy (0.00) = almost exclusively one segment. '
         'Low-entropy products are more sensitive to segment-specific shocks.'),
        ('VMS vertical concentration',
         'Government dominates Router Branch and some Switch Enterprise verticals. '
         'Manufacturing dominates Industrial products. Government = lumpy demand (budget-cycle driven). '
         'Manufacturing = more regular operational procurement.'),
        ('External data recommendation',
         'Cisco fiscal calendar is the primary seasonality driver. Additional external signals worth '
         'considering: (1) Enterprise IT budget cycle index — Q4 budget flush signal. '
         '(2) Manufacturing PMI — leading indicator for Industrial product demand. '
         '(3) Education enrollment cycles — for campus networking products. '
         '(4) Data center construction index — for DC Leaf/Spine. These are not included in the '
         'current model but can be added as lag-1 features if available.'),
        ('Correlation vs Causation note',
         'SCMS segment shares and VMS vertical shares correlate with bookings but do not cause them. '
         'They are proxy indicators for where demand originates. A spike in Government vertical '
         'share typically precedes a bookings spike 0-1 quarters later (budget approval → procurement). '
         'BigDeal Big units are concurrent with bookings spikes, not lagged — they are the same event '
         'disaggregated, so BigDeal Avg is the better forward-looking feature.'),
        ('Decline products',
         'IP Phone Enterprise Desk is in structural decline (-4.9% mean QoQ). Damped trend model '
         'prevents the model from projecting further decline beyond what the data supports. '
         'Apply a floor at 70% of most recent quarter value to prevent unrealistic low forecasts.'),
        ('NPI product — Router Core Modular Chassis',
         'Only 6 quarters of data, extremely volatile (CV=0.64), 72% big-deal driven. '
         'Own model will overfit. Use accuracy-weighted team ensemble only.'),
    ]
    for ri, (k, v) in enumerate(notes, 1):
        c1 = ws3.cell(row=ri, column=1, value=k)
        c2 = ws3.cell(row=ri, column=2, value=v)
        if ri == 1:
            c1.fill=HFILL; c1.font=HFONT; c2.fill=HFILL; c2.font=HFONT
        else:
            rf = ALT if ri % 2 == 0 else WHITE
            c1.fill=rf; c1.font=Font(name='Arial',size=9,bold=True)
            c2.fill=rf; c2.font=NFONT; c2.alignment=LFT
        c1.border=BDR; c2.border=BDR; c1.alignment=LFT
    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 110

    wb.save(OUTPUT_FILE)
    print(f'\nDone. Saved: {OUTPUT_FILE}')
    print(f'  Model_Config   : {len(config_df)} products')
    print(f'  Feature_Store  : {len(feature_store_df)} products')
    print(f'  Insights sheet : {len(notes)-1} notes')


if __name__ == '__main__':
    main()