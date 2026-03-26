"""
anomaly_detection.py
--------------------
Detects anomalous values in SCMS and VMS segment/vertical time series
using three complementary methods. Flags cells only when >= 2 methods agree.
Cross-references BigDeal data to distinguish explainable spikes from
genuine data quality issues.

Methods used
------------
1. IQR  — value exceeds Q3 + 2.5×IQR or below Q1 - 2.5×IQR per row
2. Z-score — value > 2.5 std deviations from 4-quarter rolling mean
3. Isolation Forest — sklearn anomaly detector (contamination=0.15)

Explainability check
--------------------
For each flagged cell, checks if a BigDeal order exists for that
product × quarter that covers >30% of the spike. If yes → "Monitor".
If no → "Review" (potential data quality issue).

Usage
-----
  pip install pandas openpyxl numpy scikit-learn scipy
  python anomaly_detection.py

Inputs  : SCMS_Filled.xlsx, VMS_Filled.xlsx, BigDeal.xlsx,
          Product_Description.xlsx  (all in same folder)
Output  : Anomaly_Detection_Report.xlsx  (same folder)
"""

import numpy as np
import pandas as pd
from sklearn.ensemble import IsolationForest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Config ────────────────────────────────────────────────────────────────────
SCMS_FILE    = 'SCMS_Filled.xlsx'
VMS_FILE     = 'VMS_Filled.xlsx'
BIGDEAL_FILE = 'BigDeal.xlsx'
PROD_DESC    = 'Product_Description.xlsx'
OUTPUT_FILE  = 'Anomaly_Detection_Report.xlsx'

QUARTERS = [
    '2023Q1','2023Q2','2023Q3','2023Q4',
    '2024Q1','2024Q2','2024Q3','2024Q4',
    '2025Q1','2025Q2','2025Q3','2025Q4','2026Q1',
]
BD_QUARTERS = [
    '2024Q2','2024Q3','2024Q4',
    '2025Q1','2025Q2','2025Q3','2025Q4','2026Q1',
]

IQR_MULTIPLIER       = 2.5   # tighter than standard 1.5x
ZSCORE_THRESHOLD     = 2.5   # rolling z-score cutoff
ROLLING_WINDOW       = 4     # quarters for rolling stats
ISO_CONTAMINATION    = 0.15  # fraction of anomalies for IsoForest
MIN_METHODS_TO_FLAG  = 2     # report cell only if this many methods agree
MIN_ROW_SUM          = 20    # skip near-zero rows (no signal)
BD_EXPLAIN_THRESHOLD = 0.3   # big deal covers >30% of spike → explainable


# ── Styles ────────────────────────────────────────────────────────────────────
_thin  = Side(style='thin', color='B0B0B0')
BDR    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CTR    = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
HFILL  = PatternFill('solid', start_color='1F4E79')
HFONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
NFONT  = Font(name='Arial', size=9)
ALT    = PatternFill('solid', start_color='EEF2F7')
WHITE  = PatternFill('solid', start_color='FFFFFF')
HIGH   = PatternFill('solid', start_color='FFD7D7')   # red  — HIGH severity
MED    = PatternFill('solid', start_color='FFF3CD')   # amber — MEDIUM
EXPL   = PatternFill('solid', start_color='E2EFDA')   # green — explainable
RFONT  = Font(name='Arial', size=9, bold=True, color='C00000')  # Review
MFONT  = Font(name='Arial', size=9, color='375623')             # Monitor
LEFT_COLS = {'Product','Description','Segment','Vertical','Detection_Methods'}


# ══════════════════════════════════════════════════════════════════════════════
# Detection
# ══════════════════════════════════════════════════════════════════════════════

def detect_anomalies(df, group_col, sheet_name, bd_lookup, desc_lookup):
    """
    Run multi-method anomaly detection on df.

    Parameters
    ----------
    df         : filled DataFrame (SCMS or VMS)
    group_col  : 'Segment' or 'Vertical'
    sheet_name : label for the Sheet column in results
    bd_lookup  : dict (product, quarter) → BigDeal units
    desc_lookup: dict product → description

    Returns
    -------
    DataFrame of flagged cells
    """
    results = []

    for _, row in df.iterrows():
        vals = row[QUARTERS].astype(float).values
        prod = row['Product_Name']
        grp  = row[group_col]

        if np.nansum(vals) < MIN_ROW_SUM:
            continue

        # ── IQR bounds ───────────────────────────────────────────────────────
        q1, q3 = np.percentile(vals, 25), np.percentile(vals, 75)
        iqr    = q3 - q1
        upper  = q3 + IQR_MULTIPLIER * iqr
        lower  = max(0, q1 - IQR_MULTIPLIER * iqr)

        # ── Rolling statistics ────────────────────────────────────────────────
        series = pd.Series(vals)
        roll_m = series.rolling(ROLLING_WINDOW, min_periods=2).mean()
        roll_s = series.rolling(ROLLING_WINDOW, min_periods=2).std()

        # ── Isolation Forest (per row) ────────────────────────────────────────
        valid_vals = vals[~np.isnan(vals)]
        valid_idx  = [i for i, v in enumerate(vals) if not np.isnan(v)]
        if len(valid_vals) >= 6:
            iso  = IsolationForest(contamination=ISO_CONTAMINATION, random_state=42)
            pred = iso.fit_predict(valid_vals.reshape(-1, 1))
        else:
            pred = None

        for i, (q, v) in enumerate(zip(QUARTERS, vals)):
            if np.isnan(v):
                continue

            flags = []

            # IQR
            if v > upper and v > 10:
                flags.append('IQR_high')
            elif v < lower and lower > 5:
                flags.append('IQR_low')

            # Z-score
            rm = roll_m.iloc[i]
            rs = roll_s.iloc[i]
            if rs > 0 and not np.isnan(rm):
                z = (v - rm) / rs
                if abs(z) > ZSCORE_THRESHOLD:
                    flags.append(f'Z-score {z:+.1f}')

            # Isolation Forest
            if pred is not None and i in valid_idx:
                if pred[valid_idx.index(i)] == -1:
                    flags.append('IsoForest')

            if len(flags) < MIN_METHODS_TO_FLAG:
                continue

            # Explainability check against BigDeal
            big_val     = bd_lookup.get((prod, q), 0)
            explainable = big_val > BD_EXPLAIN_THRESHOLD * v
            severity    = 'HIGH' if len(flags) >= 3 else 'MEDIUM'

            results.append({
                'Sheet':              sheet_name,
                'Product':            prod,
                'Description':        desc_lookup.get(prod, ''),
                group_col:            grp,
                'Quarter':            q,
                'Anomalous_Value':    round(v, 1),
                'IQR_Upper_Bound':    round(upper, 1),
                'Rolling_Mean':       round(float(rm), 1) if not np.isnan(float(rm)) else '',
                'Detection_Methods':  ' | '.join(flags),
                'Severity':           severity,
                'BigDeal_Units':      big_val,
                'Explainable_by_BigDeal': 'YES' if explainable else 'NO',
                'Action_Needed':      'Monitor' if explainable else 'Review',
            })

    return pd.DataFrame(results)


# ══════════════════════════════════════════════════════════════════════════════
# Excel writing
# ══════════════════════════════════════════════════════════════════════════════

def _write_detail_sheet(wb, all_anom, sheet_name, seg_col):
    ws  = wb.create_sheet(sheet_name)
    src = sheet_name.split('_')[0]   # 'SCMS' or 'VMS'
    df  = all_anom[all_anom['Sheet'] == src][[
        'Product','Description', seg_col,'Quarter',
        'Anomalous_Value','IQR_Upper_Bound','Rolling_Mean',
        'Detection_Methods','Severity','BigDeal_Units',
        'Explainable_by_BigDeal','Action_Needed',
    ]].reset_index(drop=True)

    hdrs = [c.replace('_', ' ') for c in df.columns]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill=HFILL; c.font=HFONT; c.alignment=CTR; c.border=BDR
    ws.row_dimensions[1].height = 30

    for ri, row in df.iterrows():
        er  = ri + 2
        exp = row['Explainable_by_BigDeal']
        sev = row['Severity']
        rf  = EXPL if exp == 'YES' else (HIGH if sev == 'HIGH' else MED)
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(row=er, column=ci, value=row[col] if row[col] != '' else None)
            c.fill   = rf
            c.border = BDR
            c.alignment = LFT if col in LEFT_COLS else CTR
            c.font = RFONT if col == 'Action_Needed' and row[col] == 'Review' else (
                     MFONT if col == 'Action_Needed' else NFONT)

    widths = {
        'Product':40,'Description':50,seg_col:22,'Quarter':10,
        'Anomalous_Value':15,'IQR_Upper_Bound':16,'Rolling_Mean':14,
        'Detection_Methods':34,'Severity':10,'BigDeal_Units':15,
        'Explainable_by_BigDeal':22,'Action_Needed':14,
    }
    for ci, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 15)

    n = len(df)
    ws.cell(row=n+3, column=1, value=f'Total flagged: {n}').font = Font(name='Arial', bold=True, size=9)
    ws.cell(row=n+4, column=1,
            value=f'Need review: {(df["Action_Needed"]=="Review").sum()}').font = Font(
            name='Arial', bold=True, size=9, color='C00000')
    ws.cell(row=n+5, column=1,
            value=f'Explainable by big deal: {(df["Explainable_by_BigDeal"]=="YES").sum()}').font = Font(
            name='Arial', bold=True, size=9, color='375623')


def _write_summary_sheet(wb, all_anom):
    ws = wb.create_sheet('Summary', 0)
    ws.merge_cells('A1:H1')
    t = ws['A1']
    t.value     = 'Anomaly Detection Report — SCMS & VMS'
    t.font      = Font(name='Arial', bold=True, size=13, color='1F4E79')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    stats = [
        ('', ''),
        ('Metric', 'Value'),
        ('Total anomalies',            len(all_anom)),
        ('SCMS anomalies',             len(all_anom[all_anom['Sheet']=='SCMS'])),
        ('VMS anomalies',              len(all_anom[all_anom['Sheet']=='VMS'])),
        ('HIGH severity',              (all_anom['Severity']=='HIGH').sum()),
        ('MEDIUM severity',            (all_anom['Severity']=='MEDIUM').sum()),
        ('Explainable by big deal',    (all_anom['Explainable_by_BigDeal']=='YES').sum()),
        ('Need manual review',         (all_anom['Action_Needed']=='Review').sum()),
    ]
    for ri, (k, v) in enumerate(stats, 2):
        c1 = ws.cell(row=ri, column=1, value=k)
        c2 = ws.cell(row=ri, column=2, value=v)
        if k == 'Metric':
            for c in (c1, c2): c.fill=HFILL; c.font=HFONT; c.border=BDR
        elif k:
            rf = ALT if ri % 2 == 0 else WHITE
            for c in (c1, c2):
                c.font=NFONT; c.border=BDR; c.fill=rf
            c2.font = Font(name='Arial', size=9, bold=True)

    ws.cell(row=14, column=1, value='Top 10 products by anomaly count').font = Font(
        name='Arial', bold=True, size=10, color='1F4E79')
    for ri, (prod, cnt) in enumerate(
        all_anom.groupby('Product').size().sort_values(ascending=False).head(10).items(), 15
    ):
        c1 = ws.cell(row=ri, column=1, value=prod)
        c2 = ws.cell(row=ri, column=2, value=cnt)
        rf = ALT if ri % 2 == 0 else WHITE
        for c in (c1, c2): c.font=NFONT; c.border=BDR; c.fill=rf

    ws.cell(row=14, column=4, value='Method legend').font = Font(
        name='Arial', bold=True, size=10, color='1F4E79')
    for ri, (k, v) in enumerate([
        ('IQR_high / IQR_low', f'Outside Q3 ± {IQR_MULTIPLIER}×IQR'),
        ('Z-score',             f'>{ZSCORE_THRESHOLD}σ from {ROLLING_WINDOW}-quarter rolling mean'),
        ('IsoForest',           f'Isolation Forest outlier (contamination={ISO_CONTAMINATION})'),
        ('≥2 methods',          'Cell reported only when ≥2 methods agree'),
    ], 15):
        c1 = ws.cell(row=ri, column=4, value=k); c1.font=Font(name='Arial', size=9, bold=True); c1.border=BDR
        c2 = ws.cell(row=ri, column=5, value=v); c2.font=NFONT; c2.border=BDR
        rf = ALT if ri % 2 == 0 else WHITE
        c1.fill=rf; c2.fill=rf

    ws.cell(row=21, column=4, value='Colour legend').font = Font(
        name='Arial', bold=True, size=10, color='1F4E79')
    for ri, (fl, txt) in enumerate([
        (HIGH, 'HIGH — 3+ methods agree, needs review'),
        (MED,  'MEDIUM — 2 methods agree, needs review'),
        (EXPL, 'Explainable by big deal — monitor only'),
    ], 22):
        c1 = ws.cell(row=ri, column=4, value=''); c1.fill=fl; c1.border=BDR
        c2 = ws.cell(row=ri, column=5, value=txt); c2.font=NFONT; c2.border=BDR; c2.fill=fl

    for col, w in [('A',45),('B',12),('D',32),('E',60)]:
        ws.column_dimensions[col].width = w


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print('Loading data...')
    scms      = pd.read_excel(SCMS_FILE)
    vms       = pd.read_excel(VMS_FILE)
    bd        = pd.read_excel(BIGDEAL_FILE)
    prod_desc = pd.read_excel(PROD_DESC)

    # Build lookups
    bd_lookup   = {(r['Product_Name'], q): r[f'Big_{q}']
                   for _, r in bd.iterrows() for q in BD_QUARTERS}
    desc_lookup = dict(zip(prod_desc['Masked Products'], prod_desc['Description']))

    print('Running anomaly detection...')
    scms_anom = detect_anomalies(scms, 'Segment',  'SCMS', bd_lookup, desc_lookup)
    vms_anom  = detect_anomalies(vms,  'Vertical', 'VMS',  bd_lookup, desc_lookup)
    all_anom  = pd.concat([scms_anom, vms_anom], ignore_index=True)

    from collections import Counter
    print(f'\n  SCMS: {len(scms_anom)} anomalies | '
          f'review={(scms_anom["Action_Needed"]=="Review").sum()} | '
          f'explainable={(scms_anom["Explainable_by_BigDeal"]=="YES").sum()}')
    print(f'  VMS:  {len(vms_anom)} anomalies | '
          f'review={(vms_anom["Action_Needed"]=="Review").sum()} | '
          f'explainable={(vms_anom["Explainable_by_BigDeal"]=="YES").sum()}')
    print(f'  Total: {len(all_anom)} | '
          f'HIGH={(all_anom["Severity"]=="HIGH").sum()} | '
          f'MEDIUM={(all_anom["Severity"]=="MEDIUM").sum()}')

    print('\nWriting report...')
    wb = Workbook()
    wb.remove(wb.active)
    _write_summary_sheet(wb, all_anom)
    _write_detail_sheet(wb, all_anom, 'SCMS_Anomalies', 'Segment')
    _write_detail_sheet(wb, all_anom, 'VMS_Anomalies',  'Vertical')
    wb.save(OUTPUT_FILE)
    print(f'Saved -> {OUTPUT_FILE}')


if __name__ == '__main__':
    main()