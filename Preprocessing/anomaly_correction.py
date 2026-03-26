"""
anomaly_product_type_correction.py
-----------------------------------
Cross-checks anomalous values in SCMS and VMS against product-type
plausibility rules derived from product descriptions and SCMS/VMS
category definitions.

Logic
-----
Each product belongs to one of 5 types:
  1. Campus & Enterprise Networking
  2. Data Center Infrastructure
  3. Industrial Networking
  4. Routing, Security & WAN
  5. Collaboration & Voice

Each type has a list of suspicious segments (SCMS) and suspicious
verticals (VMS) — combinations where that product type is unlikely
to generate high demand legitimately.

For every (product, segment/vertical) row that is suspicious:
  - Detect anomalous values using IQR (Q3 + 2.5×IQR threshold)
  - Clip them to the 4-quarter rolling mean of that row
  - If rolling mean is NaN (early quarters with no history), use the
    row's non-zero median instead

All corrections are logged with original value, clipped value, and reason.
Corrected cells are highlighted in orange in the output Excel files.

Usage
-----
  pip install pandas openpyxl numpy
  python anomaly_product_type_correction.py

Inputs  : SCMS_Filled.xlsx, VMS_Filled.xlsx, Product_Description.xlsx
          (all in same folder as this script)
Outputs : SCMS_Filled_Corrected.xlsx, VMS_Filled_Corrected.xlsx
"""

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Config ────────────────────────────────────────────────────────────────────
SCMS_INPUT  = 'SCMS_Filled.xlsx'
VMS_INPUT   = 'VMS_Filled.xlsx'
SCMS_OUTPUT = 'SCMS_Filled_Corrected.xlsx'
VMS_OUTPUT  = 'VMS_Filled_Corrected.xlsx'

QUARTERS = [
    '2023Q1','2023Q2','2023Q3','2023Q4',
    '2024Q1','2024Q2','2024Q3','2024Q4',
    '2025Q1','2025Q2','2025Q3','2025Q4','2026Q1',
]

IQR_MULTIPLIER = 2.5   # same threshold used in anomaly detection


# ── Product type definitions ──────────────────────────────────────────────────
# Each type specifies which SCMS segments and VMS verticals are SUSPICIOUS
# (i.e. that product type is unlikely to generate genuinely high demand there).
# Anomalous spikes in suspicious combinations are clipped to rolling mean.

PRODUCT_TYPE_MAP = {

    'Campus & Enterprise Networking': {
        # Campus switches, enterprise switches, WiFi access points
        # Primary buyers: offices, schools, hospitals, government campuses
        # Service Providers rarely deploy campus access switches directly
        'products': [
            'SWITCH Enterprise 48-Port UPOE',
            'SWITCH Enterprise 24-Port PoE+',
            'SWITCH Enterprise 24-Port UPOE',
            'SWITCH Enterprise 48-Port Fiber (Non-PoE)',
            'SWITCH Enterprise 25G Fiber',
            'SWITCH Enterprise 48-Port PoE+ with 10G Uplink',
            'SWITCH Enterprise 24-Port PoE+ Compact',
            'SWITCH Core 25G/100G Fiber',
            'SWITCH Core 100G Fiber',
            'SWITCH Core 100G/400G Fiber',
            'WIRELESS ACCESS POINT WiFi6E (Internal Antenna) Indoor',
            'WIRELESS ACCESS POINT WiFi6 (Internal Antenna) Indoor',
            'WIRELESS ACCESS POINT WiFi6E (External Antenna) Outdoor',
            'WIRELESS ACCESS POINT WiFi6 (External Antenna) Outdoor',
            'WIRELESS ACCESS POINT WiFi6 (External Antenna) Indoor',
        ],
        'suspicious_scms': ['SERVICE PROVIDER'],
        # Hospitality rarely deploys enterprise-grade campus infrastructure at scale
        # Service Provider buys DC/core equipment, not campus access switches
        'suspicious_vms': ['Hospitality/Hotels & Leisure', 'Service Provider'],
    },

    'Data Center Infrastructure': {
        # DC leaf/spine switches, core routers — for building data centers
        # SMB/Commercial businesses don't build their own DC fabrics
        # Hospitality, Retail, Transport rarely build DC infrastructure
        'products': [
            'SWITCH Data Center 25G/100G Leaf',
            'SWITCH Data Center 100G Spine',
            'ROUTER Core Modular Chassis',
        ],
        'suspicious_scms': ['SMB', 'COMMERCIAL'],
        'suspicious_vms': [
            'Hospitality/Hotels & Leisure', 'Retail',
            'Wholesale/Distribution', 'Transportation',
        ],
    },

    'Industrial Networking': {
        # Rugged switches for factories, utilities, transport OT networks
        # Financial services, hospitality, media/entertainment, healthcare
        # don't use industrial-grade rugged switches in significant volumes
        'products': [
            'SWITCH Industrial 8-Port PoE',
            'SWITCH Industrial 8-Port PoE Compact',
            'SWITCH Industrial Managed',
            'SWITCH Industrial 24-Port PoE',
        ],
        'suspicious_scms': ['SERVICE PROVIDER', 'SMB'],
        'suspicious_vms': [
            'Financial Services', 'Hospitality/Hotels & Leisure',
            'Media/Entertainment', 'Education- Public/Private',
            'Professional Services', 'Retail', 'Health Care',
        ],
    },

    'Routing, Security & WAN': {
        # Enterprise firewalls, WAN routers, branch routers
        # SMB rarely procures enterprise-grade NGFW / edge aggregation routers
        # Hospitality very rarely deploys enterprise WAN equipment at scale
        'products': [
            'ROUTER Enterprise Edge 10G/40G',
            'ROUTER Branch 4-Port PoE',
            'ROUTER Branch 8-Port PoE',
            'ROUTER Edge Aggregation Fiber',
            'SECURITY FIREWALL Next-Generation_1',
            'SECURITY FIREWALL Next-Generation_2',
        ],
        'suspicious_scms': ['SMB'],
        'suspicious_vms': ['Hospitality/Hotels & Leisure'],
    },

    'Collaboration & Voice': {
        # IP phones, conference phones — office communication
        # Service Provider, Transportation, Energy/Utilities, Media/Entertainment
        # have specialised telephony and don't bulk-buy enterprise desk phones
        'products': [
            'IP PHONE Enterprise Desk',
            'IP CONFERENCE PHONE',
        ],
        'suspicious_scms': ['SERVICE PROVIDER'],
        'suspicious_vms': [
            'Energy/Utilities', 'Transportation',
            'Service Provider', 'Media/Entertainment',
        ],
    },
}

# Flatten: product → type
PROD_TO_TYPE = {
    p: t
    for t, info in PRODUCT_TYPE_MAP.items()
    for p in info['products']
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def is_suspicious(product, seg_or_vert, key):
    """Return True if this product-segment/vertical combination is suspicious."""
    ptype = PROD_TO_TYPE.get(product)
    return bool(ptype and seg_or_vert in PRODUCT_TYPE_MAP[ptype].get(key, []))


def apply_corrections(df, prod_col, group_col, susp_key):
    """
    For every suspicious row, find IQR anomalies and clip to rolling mean.

    Parameters
    ----------
    df        : DataFrame (SCMS or VMS, quarter cols already float)
    prod_col  : 'Product_Name'
    group_col : 'Segment' or 'Vertical'
    susp_key  : 'suspicious_scms' or 'suspicious_vms'

    Returns
    -------
    df_out : corrected DataFrame
    log    : list of dicts, one per changed cell
    """
    df  = df.copy()
    log = []

    for idx, row in df.iterrows():
        prod = row[prod_col]
        grp  = row[group_col]

        if not is_suspicious(prod, grp, susp_key):
            continue

        vals   = pd.Series(row[QUARTERS].astype(float).values, index=QUARTERS)
        roll_m = vals.rolling(4, min_periods=2).mean()
        row_med = float(vals[vals > 0].median()) if (vals > 0).any() else 0.0

        q1, q3 = vals.quantile(0.25), vals.quantile(0.75)
        iqr    = q3 - q1
        upper  = q3 + IQR_MULTIPLIER * iqr

        for q in QUARTERS:
            v = float(vals[q])
            if np.isnan(v) or v <= upper:
                continue

            rm       = float(roll_m[q])
            clip_val = round(rm if not np.isnan(rm) else row_med, 4)

            df.at[idx, q] = clip_val
            log.append({
                prod_col:      prod,
                group_col:     grp,
                'Quarter':     q,
                'Original':    round(v, 4),
                'Clipped':     clip_val,
                'Method':      'rolling_mean' if not np.isnan(rm) else 'row_median',
                'Product_Type': PROD_TO_TYPE.get(prod, ''),
                'Reason':      (
                    f"{PROD_TO_TYPE.get(prod,'')} product in suspicious "
                    f"{group_col.lower()} '{grp}' — anomalous spike clipped"
                ),
            })

    return df, log


# ── Excel writing ─────────────────────────────────────────────────────────────

_thin  = Side(style='thin', color='B0B0B0')
BDR    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CTR    = Alignment(horizontal='center', vertical='center')
LFT    = Alignment(horizontal='left',   vertical='center')
HFILL  = PatternFill('solid', start_color='1F4E79')
HFONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
NFONT  = Font(name='Arial', size=9)
ALT    = PatternFill('solid', start_color='D9E1F2')
WHITE  = PatternFill('solid', start_color='FFFFFF')
CORR_FILL = PatternFill('solid', start_color='FFE0B2')            # orange
CORR_FONT = Font(name='Arial', size=9, italic=True, bold=True, color='BF360C')
LEFT_COLS = {'Product_Name', 'Segment', 'Vertical'}


def write_corrected_excel(df, log_entries, out_path, prod_col, group_col):
    df.to_excel(out_path, index=False)
    wb = load_workbook(out_path)
    ws = wb.active
    hdr = {c.value: c.column for c in ws[1] if c.value}

    # Style header
    for c in ws[1]:
        c.fill = HFILL; c.font = HFONT
        c.alignment = CTR; c.border = BDR

    # Build correction lookup
    corr_map = {(e[prod_col], e[group_col], e['Quarter']): e for e in log_entries}

    for df_idx, row in df.iterrows():
        er   = df_idx + 2
        rf   = ALT if er % 2 == 0 else WHITE
        prod = row[prod_col]
        grp  = row[group_col]

        for col in df.columns:
            ec = hdr.get(col)
            if not ec:
                continue
            cell = ws.cell(row=er, column=ec)
            val  = row[col]
            cell.value     = None if (isinstance(val, float) and np.isnan(val)) else val
            cell.border    = BDR
            cell.alignment = LFT if col in LEFT_COLS else CTR

            if corr_map.get((prod, grp, col)):
                cell.fill = CORR_FILL
                cell.font = CORR_FONT
            else:
                cell.fill = rf
                cell.font = NFONT

    # Correction log sheet
    ws_l      = wb.create_sheet('Correction_Log')
    log_hdrs  = [prod_col, group_col, 'Quarter', 'Original', 'Clipped',
                 'Method', 'Product_Type', 'Reason']
    for ci, h in enumerate(log_hdrs, 1):
        c = ws_l.cell(row=1, column=ci, value=h)
        c.fill = HFILL; c.font = HFONT; c.alignment = CTR; c.border = BDR

    for ri, entry in enumerate(log_entries, 2):
        rf = ALT if ri % 2 == 0 else WHITE
        for ci, h in enumerate(log_hdrs, 1):
            c = ws_l.cell(row=ri, column=ci, value=entry.get(h, ''))
            c.fill = rf; c.font = NFONT
            c.border = BDR; c.alignment = CTR

    for col in ws_l.columns:
        ws_l.column_dimensions[get_column_letter(col[0].column)].width = 45

    wb.save(out_path)


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print('Loading files ...')
    scms = pd.read_excel(SCMS_INPUT)
    vms  = pd.read_excel(VMS_INPUT)

    # Cast quarter columns to float to avoid dtype issues
    for q in QUARTERS:
        scms[q] = scms[q].astype(float)
        vms[q]  = vms[q].astype(float)

    print('\nApplying product-type cross-check corrections to SCMS ...')
    scms_fixed, scms_log = apply_corrections(
        scms, 'Product_Name', 'Segment', 'suspicious_scms')

    print(f'  {len(scms_log)} cells corrected')
    for c in scms_log:
        print(f"    {c['Product_Name'][:40]:40s} | {c['Segment']:18s} | "
              f"{c['Quarter']}: {c['Original']:.1f} → {c['Clipped']:.1f} [{c['Method']}]")

    print('\nApplying product-type cross-check corrections to VMS ...')
    vms_fixed, vms_log = apply_corrections(
        vms, 'Product_Name', 'Vertical', 'suspicious_vms')

    print(f'  {len(vms_log)} cells corrected')
    for c in vms_log:
        print(f"    {c['Product_Name'][:40]:40s} | {c['Vertical']:28s} | "
              f"{c['Quarter']}: {c['Original']:.1f} → {c['Clipped']:.1f} [{c['Method']}]")

    print('\nWriting corrected files ...')
    write_corrected_excel(
        scms_fixed, scms_log, SCMS_OUTPUT, 'Product_Name', 'Segment')
    write_corrected_excel(
        vms_fixed,  vms_log,  VMS_OUTPUT,  'Product_Name', 'Vertical')

    print(f'\nDone.')
    print(f'  {SCMS_OUTPUT}  — {len(scms_log)} corrections (orange italic cells)')
    print(f'  {VMS_OUTPUT}   — {len(vms_log)} corrections (orange italic cells)')
    print('  Each file has a Correction_Log sheet with full details.')


if __name__ == '__main__':
    main()